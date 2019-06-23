import click
import openpyxl
import os
import django
os.environ.setdefault("DJANGO_SETTINGS_MODULE","IPLClone.settings")
django.setup()

from IPLCloneApp.models import deliveries,matches

@click.group()
@click.pass_context
def cli(ctx):
    pass

@cli.command()
@click.argument('source_excel', nargs=1)
def from_excel_matches(source_excel):
    wb1=openpyxl.load_workbook(source_excel)
    sheet1=wb1.get_sheet_by_name('matches')
    rows = sheet1.max_row
    columns = sheet1.max_column
    for r in range(2, rows + 1):
        row_values=[]
        for c in range(1, columns + 1):
            e = sheet1.cell(row=r, column=c)
            row_values.append(e.value)
        try:
            m = matches()
            m.id = int(row_values[0])
            m.season = row_values[1]
            m.city = row_values[2]
            m.date = row_values[3]
            m.team1 = row_values[4]
            m.team2 = row_values[5]
            m.toss_winner = row_values[6]
            m.toss_decision = row_values[7]
            m.result = row_values[8]
            if row_values[9] == 0:
                m.dl_applied = False
            else:
                m.dl_applied = True
            m.winner = row_values[10]
            m.win_by_runs = row_values[11]
            m.win_by_wickets = row_values[12]
            m.player_of_match = row_values[13]
            m.venue = row_values[14]
            m.umpire1 = row_values[15]
            m.umpire2 = row_values[16]
            m.umpire3 = row_values[17]
            m.save()
            del m
        except Exception as e:
            print(row_values)
            del m
            print(e)

@cli.command()
@click.argument('source_excel', nargs=1)
def from_excel_deliveries(source_excel):
    wb1=openpyxl.load_workbook(source_excel)
    sheet1=wb1.get_sheet_by_name('deliveries')
    rows = sheet1.max_row
    columns = sheet1.max_column
    for r in range(2, rows + 1):
        row_values=[]
        for c in range(1, columns + 1):
            e = sheet1.cell(row=r, column=c)
            row_values.append(e.value)
        print(row_values)
        try:
            d=deliveries()
            d.match_id = matches.objects.get(id=row_values[0])
            d.inning =row_values[1]
            d.batting_team = row_values[2]
            d.bowling_team = row_values[3]
            d.over = row_values[4]
            d.ball = row_values[5]
            d.batsman = row_values[6]
            d.non_striker = row_values[7]
            d.bowler = row_values[8]
            d.is_super_over = row_values[9]
            d.wide_runs = row_values[10]
            d.bye_runs = row_values[11]
            d.legbye_runs = row_values[12]
            d.noball_runs = row_values[13]
            d.penalty_runs = row_values[14]
            d.batsman_runs = row_values[15]
            d.extra_runs = row_values[16]
            d.total_runs = row_values[17]
            d.player_dismissed = row_values[18]
            d.dismissal_kind = row_values[19]
            d.fielder = row_values[20]
            d.save()
            del d
        except Exception as e:
            del d
            print(e)


if __name__ == '__main__':
    cli()
    pass