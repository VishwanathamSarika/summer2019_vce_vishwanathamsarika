import click
import openpyxl
from copy import copy
from bs4 import BeautifulSoup
import os
import django
os.environ.setdefault("DJANGO_SETTINGS_MODULE","classproject.settings")
django.setup()

from onlineapp.models import College,MockTest1,Student

@click.group()
@click.pass_context
def cli(ctx):
    pass


@cli.command()
@click.argument('source_excel', nargs=1)
def from_excel(source_excel):
    wb1=openpyxl.load_workbook(source_excel)
    sheet1=wb1.get_sheet_by_name('Colleges')
    rows = sheet1.max_row
    columns = sheet1.max_column
    for r in range(2, rows + 1):
        row_values=[]
        for c in range(1, columns + 1):
            e = sheet1.cell(row=r, column=c)
            row_values.append(e.value)
        c=College()
        c.name=row_values[0]
        c.location=row_values[2]
        c.acronym=row_values[1]
        c.contact=row_values[3]
        c.save()
        del c

    sheet1 = wb1.get_sheet_by_name('Current')
    rows = sheet1.max_row
    columns = sheet1.max_column
    for r in range(2, rows + 1):
        row_values = []
        for c in range(1, columns + 1):
            e = sheet1.cell(row=r, column=c)
            row_values.append(e.value)
        try:
            s=Student()
            s.name=row_values[0]
            s.email=row_values[2]
            s.db_folder=row_values[3]
            s.college=College.objects.get(acronym=row_values[1])
            s.save()
            del s
        except Exception as e:
            del s
            print(e)
    pass

@cli.command()
@click.argument('source_html', nargs=1)
def from_html(source_html):
    r=open(source_html,"r")
    html_content=r.read()
    soup=BeautifulSoup(html_content,"html.parser")
    rows = soup.find_all('tr')
    for r in range(1,len(rows)):
        cells = rows[r].find_all(['th', 'td'])
        row_values=[]
        for c in range(1,len(cells)):
            row_values.append(cells[c].text)
        try:
            m=MockTest1()
            m.problem1=int(row_values[1])
            m.problem2=int(row_values[2])
            m.problem3=int(row_values[3])
            m.problem4=int(row_values[4])
            m.total=int(row_values[5])
            m.student=Student.objects.get(db_folder=row_values[0].split('_')[2])
            m.save()
            del m
        except Exception as e:
            del m
            print(e)
    pass

def add_dropouts():
    wb1 = openpyxl.load_workbook("students.xlsx")
    sheet1 = wb1.get_sheet_by_name('Deletions')
    rows = sheet1.max_row
    columns = sheet1.max_column
    for r in range(2, rows + 1):
        row_values = []
        for c in range(1, columns + 1):
            e = sheet1.cell(row=r, column=c)
            row_values.append(e.value)

        try:
            s = Student()
            s.name = row_values[0]
            s.email = row_values[2]
            s.db_folder = row_values[3]
            s.college = College.objects.get(acronym=row_values[1])
            s.dropped_out=True
            s.save()
            del s
        except Exception as e:
            del s
            print(e)
    pass

#add_dropouts()

if __name__ == '__main__':
    cli()
    pass